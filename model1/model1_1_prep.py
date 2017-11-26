# -*- coding: utf-8 -*-
"""
Created on Sat Nov 25 17:09:54 2017

@author: Administrator
"""
from openpyxl import load_workbook
import csv
import numpy as np


def my_float(s):
    if s is None or s == '':
        return None
    return float(s)


def my_max(a, b):
    if a is None and b is None:
        return None
    if a is None:
        return b
    elif b is None:
        return a
    elif a > b:
        return a
    else:
        return b


def model1_1_fbf_var(file_path, frame_rate, bit_rate):
    csv_reader = csv.reader(open(file_path, encoding='utf-8'))

    # parse the FFProbe data
    frame_type_list = []
    bit_size_list = []
    width = []
    height = []
    qp_min_list = []
    qp_max_list = []
    qp_list = []
    skip_ratio = []
    mv0 = []
    mv1 = []
    is_empty = True

    for row in csv_reader:
        is_empty = False
        frame_type_list.append(row[5])
        bit_size_list.append(int(row[2]))
        width.append(int(row[3]))
        height.append(int(row[4]))
        qp_min_list.append(int(row[7]))
        qp_max_list.append(int(row[8]))
        qp_list.append(float(row[9]))
        skip_ratio.append(float(row[6].strip("%")) / 100)
        mv0.append(my_float(row[12]))

        if len(row) < 16:
            mv1.append(None)
        else:
            mv1.append(my_float(row[15]))

    # default frame rate
    #  frame_rate = 25;

    # Some videos without B frames
    # to simplify the process, set mv1 as mv0
    if is_empty:
        return [0] * 11
    if len(frame_type_list) > 12:
        pass
    else:
        mv1 = mv0

    # I-frame Flicker Detection
    i_intra_flicker = 0.
    ind = [i for i, a in enumerate(frame_type_list) if a == "I"]
    for j in range(1, len(ind) - 1):
        if (qp_list[ind[j]] - qp_list[ind[j + 1]] > 5) and (qp_list[ind[j]] - qp_list[ind[j - 1]] > 5):
            i_intra_flicker = 1
            flicker_location = ind[j]
            break
        else:
            flicker_location = 0
            continue


            ##IPB ststistics

    # Temperal pooling for frame quality.
    # frame_quality = []; data_smooth = [];

    # QP list for different frame type

    # calculate the frame qulity for each frame type
    # currently only I/P frames will be counted.
    count_i = 0

    count_all = 0

    count_ii = 0
    count_pp = 0
    count_bb = 0  # caltulate the num of I/B/P
    pixels = height[0] * width[0]

    #   caltulate the num of I/B/P
    i_location = np.array(ind.copy())
    count_ii = len(i_location)
    count_pp = len([i for i, a in enumerate(frame_type_list) if a == "P"])
    count_bb = len([i for i, a in enumerate(frame_type_list) if a == "B"])
    nbr_between_two_i_pics = count_pp / count_ii
    kfr = frame_rate / nbr_between_two_i_pics

    frame_i_sts = []
    for i in range(len(qp_list)):
        frame_type = frame_type_list[i]
        if frame_type == 'I' and qp_list[i] > 2:
            count_i = count_i + 1
            count_all = count_all + 1
            frame_i_sts.append(bit_size_list[i])

    ans = [np.mean(qp_list), frame_rate, i_intra_flicker, np.mean(qp_max_list), np.mean(qp_min_list), bit_rate,
           np.mean(frame_i_sts), np.mean(skip_ratio), np.mean([my_max(mv0[i], mv1[i]) for i, a in enumerate(mv0)]), kfr,
           width[0] * height[0]]

    return ans


def batch():
    xdata = []
    # csvFile = open("instance.csv", "w",newline='')#xdata is stored here
    # writer = csv.writer(csvFile)
    wb = load_workbook(filename="C:\\Users\\Administrator\\Desktop\\my python\\U-vMOS VBR盖亚_TV_v0.91_SJTU.xlsx")
    ws = wb.get_sheet_by_name('video_fr_br')
    for n in range(1, 71):
        csv_filepath = "C:\\Users\\Administrator\\Desktop\\model1资料\\资料\\Ori_data\\ffprobe_csv_huawei"
        frame_rate = ws.cell(row=n + 1, column=6).value
        bit_rate = ws.cell(row=n + 1, column=7).value

        tmp = model1_1_fbf_var(csv_filepath + '\\' + str(n) + '.csv', frame_rate, bit_rate)
        # writer.writerow(l,)
        xdata.append(tmp)
    # csvFile.close()
    return xdata


if __name__ == '__main__':
    print(batch())
