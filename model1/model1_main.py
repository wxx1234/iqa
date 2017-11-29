from model1.model1_1 import model1_1
from model1.model1_2 import model1_2
import pandas as pd
from openpyxl import load_workbook
import time


def model1(filename, frame_rate, bit_rate):
    csv_data = pd.read_csv(filename, names=['frame', 'pkt_pts', 'pkt_size', 'width', 'height', 'pict_type',
                                            'skip_ratio', 'qp_min', 'qp_max', 'qp_avg', 'mv0_min', 'mv0_max',
                                            'mv0_avg', 'mv1_min', 'mv1_max', 'mv1_avg'])

    width = csv_data['width']
    height = csv_data['height']
    qp_list = csv_data['qp_avg']
    frame_type_list = csv_data['pict_type']
    skip_ratio = csv_data['skip_ratio']
    for i in range(len(skip_ratio)):
        skip_ratio.set_value(i, float(skip_ratio[i].strip("%")) / 100)
    mv0 = csv_data['mv0_avg']
    csv_data = csv_data.dropna(axis=1, how='all')
    number_of_frames, number_of_params = csv_data.shape
    if number_of_params > 13:
        mv1 = csv_data['mv1_avg']
    else:
        mv1 = mv0
    number_of_frames, number_of_params = csv_data.shape
    bit_size_list = csv_data['pkt_size']
    qp_min_list = csv_data['qp_min']
    qp_max_list = csv_data['qp_max']
    model1_1_mos = model1_1(width=width, height=height, qp_list=qp_list, frame_type_list=frame_type_list,
                            skip_ratio=skip_ratio, mv0=mv0, mv1=mv1, bit_size_list=bit_size_list,
                            qp_max_list=qp_max_list, qp_min_list=qp_min_list, frame_rate=frame_rate, bit_rate=bit_rate)
    model1_2_mos = model1_2(width=width, height=height, qp_list=qp_list, frame_type_list=frame_type_list,
                            skip_ratio=skip_ratio, mv0=mv0, mv1=mv1, number_of_frames=number_of_frames,
                            number_of_params=number_of_params, device='TV')
    mos_all = model1_2_mos - (5 - model1_1_mos) * (model1_2_mos - 4) / 100
    return mos_all


if __name__ == '__main__':
    wb = load_workbook(filename="C:/Users/WXX/Desktop/model1_tool/Ori_data/MOS_fr_br.xlsx")
    ws = wb.get_sheet_by_name('Sheet1')
    t = time.time()
    for number in range(1, 116):
        try:
            br = ws.cell(row=number + 1, column=2).value
            fr = ws.cell(row=number + 1, column=3).value
            mos = model1(f'C:/Users/WXX/Desktop/model1_tool/Ori_data/sequence_info/{number}.csv', frame_rate=fr,
                         bit_rate=br)
            print(number, mos)
        except:
            print(f'---process {number}.csv fail---')
    t = time.time() - t
    print(f'total {t} seconds, averge {t/115} seconds')
